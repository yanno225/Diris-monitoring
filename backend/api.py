"""
Econersys Afrique — API Backend
Monitoring Socomec Diris A40

Routes :
  - /auth/login          : connexion (retourne un token JWT)
  - /auth/register       : créer un utilisateur (admin seulement)
  - /auth/me             : infos de l'utilisateur connecté
  - /mesure/live          : dernière mesure en temps réel
  - /historique/{champ}   : historique avec filtre période
  - /stats/{champ}        : statistiques min/max/moyenne
  - /export/excel         : télécharger les données en Excel (admin seulement)
  - /ws                   : WebSocket temps réel
"""

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, Depends, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from contextlib import asynccontextmanager
import asyncio
import json
import time
import os
import sys
import io
from datetime import datetime, timezone

sys.path.append(os.path.dirname(__file__))

from database import GestionnaireDB
from auth import (
    authentifier_utilisateur,
    creer_utilisateur,
    creer_token,
    get_utilisateur_courant,
    exiger_admin,
)
from dotenv import load_dotenv

load_dotenv()

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

SIMULATEUR = os.getenv("SIMULATEUR", "false").lower() == "true"
INTERVALLE_LECTURE = 5

db = GestionnaireDB()
clients_ws = []
derniere_mesure = {}

# ─────────────────────────────────────────────
# MODÈLES PYDANTIC (validation des requêtes)
# ─────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str
    mot_de_passe: str

class RegisterRequest(BaseModel):
    email: str
    mot_de_passe: str
    nom: str
    role: str = "operateur"

# ─────────────────────────────────────────────
# TÂCHE DE LECTURE (mode bureau — lit Atlas)
# ─────────────────────────────────────────────

async def tache_lecture_compteur():
    """
    Mode BUREAU : on ne lit plus le compteur directement.
    Le Raspberry Pi s'en charge et envoie vers Atlas.
    On lit Atlas pour mettre à jour les clients WebSocket.
    """
    global derniere_mesure

    print(f"\n  ▶ Mode BUREAU — lecture depuis Atlas uniquement")
    print(f"  ▶ Le Raspberry Pi collecte les données sur le terrain")
    print(f"  ▶ Intervalle de vérification : {INTERVALLE_LECTURE}s\n")

    while True:
        try:
            mesure = db.get_derniere_mesure()

            if mesure and mesure != derniere_mesure:
                derniere_mesure = mesure

                if clients_ws:
                    message = json.dumps({
                        "type": "mesure",
                        "donnees": mesure
                    }, default=str)

                    clients_a_supprimer = []
                    for client in list(clients_ws):
                        try:
                            await client.send_text(message)
                        except:
                            clients_a_supprimer.append(client)
                    for client in clients_a_supprimer:
                        if client in clients_ws:
                            clients_ws.remove(client)

        except Exception as e:
            print(f"  ✗ Erreur tâche lecture : {e}")

        await asyncio.sleep(INTERVALLE_LECTURE)

# ─────────────────────────────────────────────
# DÉMARRAGE ET ARRÊT
# ─────────────────────────────────────────────

async def lifespan(app: FastAPI):
    print("\n" + "=" * 50)
    print("  Démarrage du serveur")
    print("=" * 50)

    if not db.connecter():
        print("  ✗ ERREUR CRITIQUE : impossible de se connecter à Atlas")
    else:
        asyncio.create_task(tache_lecture_compteur())

    yield

    print("\n  Arrêt du serveur...")
    db.fermer()
    print("  ✓ Serveur arrêté proprement.")

# ─────────────────────────────────────────────
# CRÉATION APP FASTAPI
# ─────────────────────────────────────────────

app = FastAPI(
    title="API Diris A40 - Econersys Afrique",
    description="API temps réel pour la surveillance électrique",
    version="2.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────
# ROUTES AUTHENTIFICATION
# ─────────────────────────────────────────────

@app.post("/auth/login")
async def login(requete: LoginRequest):
    """
    Connexion d'un utilisateur.
    Retourne un token JWT si les identifiants sont valides.

    Body JSON :
    {
        "email": "admin@econersys.com",
        "mot_de_passe": "monmotdepasse"
    }
    """
    utilisateur = authentifier_utilisateur(requete.email, requete.mot_de_passe)

    if not utilisateur:
        raise HTTPException(
            status_code=401,
            detail="Email ou mot de passe incorrect"
        )

    token = creer_token({
        "email": utilisateur["email"],
        "role": utilisateur["role"],
        "nom": utilisateur["nom"],
    })

    return {
        "token": token,
        "utilisateur": utilisateur,
        "message": "Connexion réussie"
    }

@app.post("/auth/register")
async def register(requete: RegisterRequest, utilisateur=Depends(exiger_admin)):
    """
    Créer un nouvel utilisateur (réservé aux administrateurs).

    Body JSON :
    {
        "email": "operateur1@econersys.com",
        "mot_de_passe": "sonmotdepasse",
        "nom": "Jean Kouassi",
        "role": "operateur"
    }
    """
    nouveau = creer_utilisateur(
        email=requete.email,
        mot_de_passe=requete.mot_de_passe,
        nom=requete.nom,
        role=requete.role,
    )

    if nouveau is None:
        raise HTTPException(
            status_code=400,
            detail=f"L'email '{requete.email}' existe déjà"
        )

    return {
        "utilisateur": nouveau,
        "message": "Utilisateur créé avec succès"
    }

@app.get("/auth/me")
async def get_me(utilisateur=Depends(get_utilisateur_courant)):
    """
    Retourne les informations de l'utilisateur connecté.
    Nécessite un token JWT valide dans le header Authorization.
    """
    return utilisateur

# ─────────────────────────────────────────────
# ROUTE ACCUEIL
# ─────────────────────────────────────────────

@app.get("/")
async def accueil():
    return {
        "application": "Econersys — Monitoring Diris A40",
        "statut": "en ligne",
        "version": "2.0.0",
        "mongodb": "connecté" if db.connecte else "déconnecté",
        "clients_ws": len(clients_ws),
    }

# ─────────────────────────────────────────────
# ROUTES DONNÉES (protégées par authentification)
# ─────────────────────────────────────────────

@app.get("/mesure/live")
async def get_mesure_live(utilisateur=Depends(get_utilisateur_courant)):
    """Retourne la dernière mesure lue."""
    if not db.connecte:
        raise HTTPException(status_code=503, detail="Base de données non connectée")

    mesure = db.get_derniere_mesure()

    if not mesure:
        if derniere_mesure:
            return derniere_mesure
        else:
            raise HTTPException(status_code=404, detail="Aucune mesure disponible")

    return mesure

@app.get("/historique/{champ}")
async def get_historique(
    champ: str,
    heures: int = 24,
    limite: int = 500,
    debut: float = None,
    fin: float = None,
    utilisateur=Depends(get_utilisateur_courant),
):
    """
    Retourne l'évolution d'un paramètre sur une période.

    Paramètres URL :
    - champ  : le paramètre (ex: cos_phi, tensions.u12, puissances.active)
    - heures : nombre d'heures (défaut: 24) — ignoré si debut/fin sont fournis
    - limite : nombre max de points (défaut: 500)
    - debut  : timestamp UNIX de début (optionnel)
    - fin    : timestamp UNIX de fin (optionnel)

    Exemples :
      /historique/cos_phi
      /historique/tensions.u12?heures=48
      /historique/puissances.active?debut=1712345678&fin=1712432078
    """
    if not db.connecte:
        raise HTTPException(status_code=503, detail="Base de données non connectée")

    champ = champ.replace("%2E", ".")

    # Si début et fin sont fournis, on utilise la méthode par période
    if debut is not None and fin is not None:
        donnees = db.get_historique_periode(champ, debut, fin, limite)
    else:
        donnees = db.get_historique(champ, periodes_heures=heures, limite=limite)

    return {
        "champ": champ,
        "heures": heures,
        "debut": debut,
        "fin": fin,
        "nb_points": len(donnees),
        "donnees": donnees,
    }

@app.get("/stats/{champ}")
async def get_stats(
    champ: str,
    heures: int = 24,
    debut: float = None,
    fin: float = None,
    utilisateur=Depends(get_utilisateur_courant),
):
    """
    Retourne Min, Max et Moyenne d'un paramètre.
    Supporte aussi les timestamps début/fin.
    """
    if not db.connecte:
        raise HTTPException(status_code=503, detail="Base de données non connectée")

    champ = champ.replace("%2E", ".")

    if debut is not None and fin is not None:
        stats = db.get_stats_periode(champ, debut, fin)
    else:
        stats = db.get_stats(champ, periode_heures=heures)

    return {
        "champ": champ,
        "heures": heures,
        "debut": debut,
        "fin": fin,
        **stats,
    }

# ─────────────────────────────────────────────
# EXPORT EXCEL (admin seulement)
# ─────────────────────────────────────────────

@app.get("/export/excel")
async def export_excel(
    debut: float = None,
    fin: float = None,
    heures: int = 24,
    utilisateur=Depends(exiger_admin),
):
    """
    Télécharge toutes les données en fichier Excel.
    Réservé aux administrateurs.

    Paramètres :
    - debut  : timestamp UNIX de début (optionnel)
    - fin    : timestamp UNIX de fin (optionnel)
    - heures : nombre d'heures (défaut: 24, ignoré si debut/fin fournis)

    Exemples :
      /export/excel
      /export/excel?heures=48
      /export/excel?debut=1712345678&fin=1712432078
    """
    if not db.connecte:
        raise HTTPException(status_code=503, detail="Base de données non connectée")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Récupérer les données de MongoDB
        if debut is not None and fin is not None:
            ts_debut = debut
            ts_fin = fin
        else:
            ts_fin = time.time()
            ts_debut = ts_fin - (heures * 3600)

        filtre = {"timestamp": {"$gte": ts_debut, "$lte": ts_fin}}

        curseur = db.collection.find(
            filtre,
            sort=[("timestamp", 1)],
        )

        documents = list(curseur)

        if not documents:
            raise HTTPException(status_code=404, detail="Aucune donnée sur cette période")

        # Créer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Mesures Diris A40"

        # Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # En-têtes
        headers = [
            "Date / Heure",
            "U12 (V)", "U23 (V)", "U31 (V)",
            "V1 (V)", "V2 (V)", "V3 (V)",
            "I1 (A)", "I2 (A)", "I3 (A)", "In (A)",
            "P Active (kW)", "P Réactive (kVAR)", "P Apparente (kVA)",
            "Cos Phi",
            "Fréquence (Hz)",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Données
        for row_idx, doc in enumerate(documents, 2):
            tensions = doc.get("tensions", {})
            courants = doc.get("courants", {})
            puissances = doc.get("puissances", {})

            valeurs = [
                doc.get("heure_abidjan", ""),
                tensions.get("u12"),
                tensions.get("u23"),
                tensions.get("u31"),
                tensions.get("v1"),
                tensions.get("v2"),
                tensions.get("v3"),
                courants.get("i1"),
                courants.get("i2"),
                courants.get("i3"),
                courants.get("in"),
                puissances.get("active"),
                puissances.get("reactive"),
                puissances.get("apparente"),
                doc.get("cos_phi"),
                doc.get("frequence"),
            ]

            for col, valeur in enumerate(valeurs, 1):
                cell = ws.cell(row=row_idx, column=col, value=valeur)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Ajuster la largeur des colonnes
        ws.column_dimensions["A"].width = 22
        for col_letter in ["B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = 12
        for col_letter in ["H", "I", "J", "K"]:
            ws.column_dimensions[col_letter].width = 12
        for col_letter in ["L", "M", "N"]:
            ws.column_dimensions[col_letter].width = 16
        ws.column_dimensions["O"].width = 12
        ws.column_dimensions["P"].width = 16

        # Figer la première ligne
        ws.freeze_panes = "A2"

        # Feuille résumé
        ws_resume = wb.create_sheet("Résumé")
        ws_resume.cell(row=1, column=1, value="Rapport Econersys Afrique").font = Font(size=16, bold=True)
        ws_resume.cell(row=2, column=1, value=f"Exporté par : {utilisateur['nom']} ({utilisateur['email']})")
        ws_resume.cell(row=3, column=1, value=f"Date d'export : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_resume.cell(row=4, column=1, value=f"Nombre de mesures : {len(documents)}")

        if documents:
            ws_resume.cell(row=5, column=1, value=f"Première mesure : {documents[0].get('heure_abidjan', '')}")
            ws_resume.cell(row=6, column=1, value=f"Dernière mesure : {documents[-1].get('heure_abidjan', '')}")

        ws_resume.column_dimensions["A"].width = 60

        # Sauvegarder dans un buffer mémoire
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Nom du fichier avec la date
        nom_fichier = f"econersys_mesures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nom_fichier}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"  ✗ Erreur export Excel : {e}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération Excel : {str(e)}")

# ─────────────────────────────────────────────
# WEBSOCKET — Temps réel
# ─────────────────────────────────────────────

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    clients_ws.append(websocket)
    print(f"  ✓ Client WebSocket connecté ({len(clients_ws)} total)")

    try:
        if derniere_mesure:
            await websocket.send_text(json.dumps({
                "type": "mesure",
                "donnees": derniere_mesure
            }, default=str))

        while True:
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text(json.dumps({"type": "pong"}))

    except WebSocketDisconnect:
        if websocket in clients_ws:
            clients_ws.remove(websocket)
        print(f"  Client WebSocket déconnecté ({len(clients_ws)} restants)")

    except Exception as e:
        if websocket in clients_ws:
            clients_ws.remove(websocket)
        print(f"  ✗ Erreur WebSocket : {e}")

# ─────────────────────────────────────────────
# LANCEMENT DU SERVEUR
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    print("\n  Lancement du serveur sur http://localhost:8080")
    print("  Documentation API : http://localhost:8080/docs")
    print("  Arrêt : Ctrl+C\n")
    uvicorn.run("api:app", host="0.0.0.0", port=8080, reload=True)





